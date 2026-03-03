#!/usr/bin/env python3
"""
npm-audit.json → Excel Remediation Tracker
===========================================
Usage:
    --- Runs looking for file named npm-audit.json in root folder, will generate generic named output .xlsx file
    python npm_audit_to_excel.py

    --- Runs looking for JSON file in root matching name provided as first argument, will generate generic named output .xlsx file 
    python npm_audit_to_excel.py <name_of_your_npm_audit>.json

    --- Runs looking for JSON file in root matching name provided as second argument, will generate named .xlsx file matching second argument  
    python npm_audit_to_excel.py <name_of_your_npm_audit>.json <name_of_excel_file_to_export_to>.xlsx

    --- Same as above, but passing in a package.json as the third argument allows for identification of dev dependencies in output .xlsx file
    python npm_audit_to_excel.py <name_of_your_npm_audit>.json <name_of_excel_file_to_export_to>.xlsx package.json

    
    Example Execution: 
    python npm_audit_to_excel.py npm-audit.json mdh-vulnerabilities-workbook.xlsx package.json

Requirements:
    pip install openpyxl

-------------------------------------------------------------------------------
NOTE ON ROW COUNT vs. `npm audit` SUMMARY
-------------------------------------------------------------------------------
Running `npm audit` on the command line reports a count of VULNERABLE PACKAGES
(e.g. "189 vulnerabilities (7 low, 114 moderate, 51 high, 17 critical)").

This script will almost always show a HIGHER row count because each package is
expanded by its "via" array — one row per individual CVE/advisory rather than
one row per package.

Example: a single package with 3 CVEs in its "via" array becomes 3 rows here.
This is intentional: it lets you track and resolve each vulnerability
individually rather than treating a multi-CVE package as a single checkbox.

If a package appears on multiple rows, ALL rows must be marked "Fixed" before
that package is truly remediated.

-------------------------------------------------------------------------------
NOTE ON DEV DEPENDENCY DETECTION (requires package.json)
-------------------------------------------------------------------------------

npm audit JSON does not include a field indicating whether a vulnerable package
is a dev or production dependency. This matters because dev dependencies do not
ship to production and are generally lower remediation priority.

To populate the "Dev Dep?" column, pass your package.json as the third argument.
The script cross-references the top-level "devDependencies" block against each
vulnerable package name and marks it "Yes", "No", or "Unknown".

Important caveats:
  - Only TOP-LEVEL devDependencies are checked. A transitive dependency
    (isDirect: No) may be pulled in by either a dev or prod package — this
    script cannot trace that chain without a full dependency tree
    (use `npm ls --json` for that level of detail).
  - If no package.json is provided, "Dev Dep?" shows "Unknown".

-------------------------------------------------------------------------------
NOTE ON THE "Done" CHECKBOX COLUMN
-------------------------------------------------------------------------------
Excel does not support native checkboxes via openpyxl. The "Done" column is an
empty cell — type ✓ (or any character) to mark a row complete. Use alongside
the Status dropdown for full tracking:
  - Done column:  quick visual tick once a row is fully resolved
  - Status:       Open / In Progress / Fixed / Won't Fix
-------------------------------------------------------------------------------
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SEVERITY_ORDER = ["critical", "high", "moderate", "low", "info"]

SEVERITY_COLORS = {
    "critical": {"header_bg": "7B0000", "header_fg": "FFFFFF", "row_bg": "FFE0E0"},
    "high":     {"header_bg": "C0392B", "header_fg": "FFFFFF", "row_bg": "FDECEA"},
    "moderate": {"header_bg": "E67E22", "header_fg": "FFFFFF", "row_bg": "FEF9E7"},
    "low":      {"header_bg": "2980B9", "header_fg": "FFFFFF", "row_bg": "EBF5FB"},
    "info":     {"header_bg": "27AE60", "header_fg": "FFFFFF", "row_bg": "EAFAF1"},
}

STATUS_OPTIONS = ["Open", "In Progress", "Fixed", "Won't Fix"]
STATUS_COLORS = {
    "Open":        {"bg": "FFE0E0", "fg": "7B0000"},
    "In Progress": {"bg": "FEF9E7", "fg": "7D5A00"},
    "Fixed":       {"bg": "E9F7EF", "fg": "1D6A39"},
    "Won't Fix":   {"bg": "F2F3F4", "fg": "555555"},
}

# Col 1: Done (tick)  Col 2: Status dropdown  Col 6: Dev Dep? (from package.json)
COLUMNS = [
    ("Done",           6),
    ("Status",         14),
    ("Package",        28),
    ("Severity",       12),
    ("Direct Dep?",    12),
    ("Dev Dep?",       10),
    ("Vulnerability",  52),
    ("Advisory URL",   50),
    ("Affected Range", 18),
    ("Fix Range",      18),
    ("Fix Available",  14),
    ("Nodes",          40),
    ("Notes",          30),
]

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── package.json ───────────────────────────────────────────────────────────────

def load_dev_deps(path: str) -> set:
    try:
        with open(path, "r", encoding="utf-8") as f:
            pkg = json.load(f)
        return set(pkg.get("devDependencies", {}).keys())
    except Exception as e:
        print(
            f"Warning: could not read package.json ({e}). Dev Dep? will show 'Unknown'.")
        return set()


# ── Parsing ────────────────────────────────────────────────────────────────────

def _fix_label(f):
    if isinstance(f, bool):
        return "Yes" if f else "No"
    if isinstance(f, dict):
        return f.get("name", "Yes")
    return str(f)


def parse_vulnerabilities(audit_data: dict, dev_deps: set) -> dict:
    groups = {s: [] for s in SEVERITY_ORDER}
    source = audit_data.get(
        "vulnerabilities", audit_data.get("advisories", {}))

    for pkg, d in source.items():
        sev = d.get("severity", "info").lower()
        is_dev = ("Yes" if pkg in dev_deps else "No") if dev_deps else "Unknown"
        base = {
            "package":       pkg,
            "severity":      sev,
            "is_direct":     "Yes" if d.get("isDirect") else "No",
            "is_dev":        is_dev,
            "fix_range":     d.get("range", ""),
            "fix_available": _fix_label(d.get("fixAvailable", False)),
            "nodes":         ", ".join(d.get("nodes", [])),
        }
        advs = [v for v in d.get("via", []) if isinstance(v, dict)]
        if advs:
            for a in advs:
                groups.setdefault(sev, []).append(
                    {**base, "title": a.get("title", ""), "url": a.get("url", ""), "adv_range": a.get("range", "")})
        else:
            strs = [v for v in d.get("via", []) if isinstance(v, str)]
            groups.setdefault(sev, []).append(
                {**base, "title": ", ".join(strs), "url": "", "adv_range": ""})

    return {s: r for s, r in groups.items() if r}


# ── Styling ────────────────────────────────────────────────────────────────────

def _hdr(cell, bg, fg):
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _dat(cell, bg, wrap=False, center=False, bold=False, color="000000"):
    cell.font = Font(name="Arial", size=9, bold=bold, color=color)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap,
                               horizontal="center" if center else "left")
    cell.border = BORDER


def _status_cell(cell, status):
    c = STATUS_COLORS.get(status, STATUS_COLORS["Open"])
    cell.value = status
    cell.font = Font(name="Arial", bold=True, size=9, color=c["fg"])
    cell.fill = PatternFill("solid", start_color=c["bg"])
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = BORDER


def _done_cell(cell, bg):
    cell.value = ""
    cell.font = Font(name="Arial", size=11, color="1D6A39", bold=True)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = BORDER


def _dev_dep_cell(cell, is_dev):
    color_map = {
        "Yes":     ("FFF3CD", "856404"),  # amber  — dev only, lower priority
        "No":      ("E9F7EF", "1D6A39"),  # green  — prod, higher priority
        "Unknown": ("F2F3F4", "555555"),  # grey
    }
    bg, fg = color_map.get(is_dev, color_map["Unknown"])
    cell.value = is_dev
    cell.font = Font(name="Arial", size=9, bold=True, color=fg)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="top")
    cell.border = BORDER


# ── Sheet builders ─────────────────────────────────────────────────────────────

def write_severity_sheet(wb: Workbook, severity: str, rows: list) -> int:
    colors = SEVERITY_COLORS.get(severity, SEVERITY_COLORS["info"])
    ws = wb.create_sheet(title=severity.capitalize())

    dv = DataValidation(type="list",
                        formula1='"' + ",".join(STATUS_OPTIONS) + '"',
                        allow_blank=False, showDropDown=False)
    dv.sqref = "B2:B10000"
    ws.add_data_validation(dv)

    ws.row_dimensions[1].height = 30
    for i, (name, width) in enumerate(COLUMNS, 1):
        _hdr(ws.cell(1, i, name), colors["header_bg"], colors["header_fg"])
        ws.column_dimensions[get_column_letter(i)].width = width

    for ri, row in enumerate(rows, 2):
        bg = colors["row_bg"] if ri % 2 == 0 else "FFFFFF"

        # Done
        _done_cell(ws.cell(ri, 1), bg)
        # Status
        _status_cell(ws.cell(ri, 2), "Open")
        _dat(ws.cell(ri, 3, row["package"]),
             bg)                          # Package
        _dat(ws.cell(ri, 4, row["severity"].capitalize()),
             bg, center=True)  # Severity
        _dat(ws.cell(ri, 5, row["is_direct"]), bg,
             center=True)           # Direct Dep?
        # Dev Dep?
        _dev_dep_cell(ws.cell(ri, 6), row["is_dev"])
        _dat(ws.cell(ri, 7, row["title"]), bg,
             wrap=True)                 # Vulnerability
        # Advisory URL
        _dat(ws.cell(ri, 8, row["url"]), bg)
        _dat(ws.cell(ri, 9, row["adv_range"]), bg,
             center=True)           # Affected Range
        _dat(ws.cell(ri, 10, row["fix_range"]),
             bg, center=True)          # Fix Range
        _dat(ws.cell(ri, 11, row["fix_available"]),
             bg, center=True)      # Fix Available
        _dat(ws.cell(ri, 12, row["nodes"]), bg,
             wrap=True)                # Nodes
        _dat(ws.cell(ri, 13, ""), bg, wrap=True)  # Notes

        ws.row_dimensions[ri].height = 40

    ws.freeze_panes = "C2"  # freeze Done + Status while scrolling right
    ws.auto_filter.ref = ws.dimensions
    return len(rows)


def write_summary_sheet(wb: Workbook, counts: dict, audit_meta: dict, has_package_json: bool):
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "npm audit – Vulnerability Remediation Tracker"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    fields = [
        ("Audit date",   audit_meta.get("created", "N/A")),
        ("npm version",  audit_meta.get("npmVersion", "N/A")),
        ("Node version", audit_meta.get("nodeVersion", "N/A")),
        ("Dependencies", str(audit_meta.get("totalDependencies", "N/A"))),
        ("package.json", "Loaded — Dev Dep? populated" if has_package_json
                         else "Not provided — Dev Dep? shows 'Unknown'"),
    ]
    for i, (l, v) in enumerate(fields, 2):
        ws.cell(i, 1, l).font = Font(name="Arial", bold=True, size=9)
        ws.cell(i, 2, v).font = Font(name="Arial", size=9)

    def _note(row, text, bg, border_color):
        ws.merge_cells(f"A{row}:E{row}")
        n = ws[f"A{row}"]
        n.value = text
        n.font = Font(name="Arial", italic=True, size=8, color="555555")
        n.fill = PatternFill("solid", start_color=bg)
        n.alignment = Alignment(wrap_text=True, vertical="center")
        s = Side(style="thin", color=border_color)
        n.border = Border(left=s, right=s, top=s, bottom=s)
        ws.row_dimensions[row].height = 45

    note_row = 2 + len(fields) + 1
    _note(note_row,
          "⚠  Row count exceeds npm audit's reported count. Each row = one CVE/advisory. "
          "A package with multiple CVEs appears on multiple rows — ALL must be marked Fixed "
          "to fully remediate it.",
          "FEF9E7", "E67E22")
    _note(note_row + 1,
          "ℹ  Dev Dep? is cross-referenced from top-level devDependencies in package.json. "
          "Transitive deps (Direct Dep? = No) may be pulled in by either dev or prod packages "
          "— this column reflects direct top-level classification only. Dev deps do not ship "
          "to production and are generally lower remediation priority.",
          "EBF5FB", "2980B9")

    start = note_row + 3
    ws.merge_cells(f"A{start}:E{start}")
    h = ws[f"A{start}"]
    h.value = "Vulnerabilities by Severity"
    h.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    h.fill = PatternFill("solid", start_color="2C3E50")
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start].height = 24
    start += 1

    for ci, hd in enumerate(["Severity", "CVE Rows", "Sheet"], 1):
        c2 = ws.cell(start, ci, hd)
        c2.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c2.fill = PatternFill("solid", start_color="566573")
        c2.alignment = Alignment(horizontal="center")
    start += 1

    total = 0
    for sev in SEVERITY_ORDER:
        n = counts.get(sev, 0)
        if not n:
            continue
        co = SEVERITY_COLORS[sev]
        for col in range(1, 4):
            ws.cell(start, col).fill = PatternFill(
                "solid", start_color=co["row_bg"])
            ws.cell(start, col).border = BORDER
        ws.cell(start, 1, sev.capitalize()).font = Font(
            name="Arial", bold=True, size=9)
        ws.cell(start, 2, n).font = Font(name="Arial", size=9)
        ws.cell(start, 2).alignment = Alignment(horizontal="center")
        ws.cell(start, 3, sev.capitalize()).font = Font(name="Arial", size=9,
                                                        color="0000EE", underline="single")
        total += n
        start += 1

    for col in range(1, 4):
        ws.cell(start, col).border = BORDER
        ws.cell(start, col).fill = PatternFill("solid", start_color="D5D8DC")
    ws.cell(start, 1, "TOTAL").font = Font(name="Arial", bold=True, size=9)
    ws.cell(start, 2, total).font = Font(name="Arial", bold=True, size=9)
    ws.cell(start, 2).alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDE", [16, 30, 12, 14, 20]):
        ws.column_dimensions[col].width = w


# ── Entry point ────────────────────────────────────────────────────────────────

def convert(input_path: str, output_path: str, package_json_path: str = None):
    with open(input_path) as f:
        audit_data = json.load(f)

    dev_deps = load_dev_deps(package_json_path) if package_json_path else set()
    has_package_json = bool(
        package_json_path and Path(package_json_path).exists())

    groups = parse_vulnerabilities(audit_data, dev_deps)
    if not groups:
        print("No vulnerabilities found.")
        return

    wb = Workbook()
    meta = audit_data.get("metadata", {})
    counts = {}
    for sev in SEVERITY_ORDER:
        if rows := groups.get(sev):
            counts[sev] = write_severity_sheet(wb, sev, rows)

    write_summary_sheet(wb, counts, meta, has_package_json)
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))
    wb.save(output_path)

    total = sum(counts.values())
    print(f"Saved          : {output_path}")
    print(
        f"Total CVE rows : {total}  (exceeds npm audit count — see Summary sheet)")
    print(
        f"package.json   : {'loaded — Dev Dep? populated' if has_package_json else 'not provided — Dev Dep? shows Unknown'}")
    for s in SEVERITY_ORDER:
        if s in counts:
            print(f"  {s.capitalize():>10}: {counts[s]} rows")


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "npm-audit.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "npm-audit-report.xlsx"
    pkgj = sys.argv[3] if len(sys.argv) > 3 else None

    if not Path(inp).exists():
        print(f"Error: {inp} not found")
        sys.exit(1)
    if pkgj and not Path(pkgj).exists():
        print(
            f"Warning: package.json not found at '{pkgj}' — Dev Dep? will show 'Unknown'")
        pkgj = None

    convert(inp, out, pkgj)
